"""result_exporter — 计算结果导出为 JSON、Excel 和 CSV。

输出：
- JSON: 结构化完整结果，每个 plan_id 下是员工列表
- Excel: 每个 plan_id 一个 sheet，列 = employee_id / eligible / amount / detail
- CSV: 薪酬提交格式，扁平宽表（每员工一行，子项横向展开）

文件命名: result_{YYYYMM}_{plan_type}.json / .xlsx / .csv
默认输出到 salary_performance_skills/output/
"""

import csv
import json
import logging
from pathlib import Path

import openpyxl

from employee_directory import EmployeeDirectory

logger = logging.getLogger("incentive_bot")


def _fmt_amount(x) -> str:
    """格式化金额为2位小数字符串。空值返回空串。"""
    if x is None or x == "":
        return ""
    try:
        return f"{float(x):.2f}"
    except (TypeError, ValueError):
        return str(x)


class ResultExporter:
    """导出计算结果到 JSON + Excel 文件。"""

    def __init__(
        self,
        output_dir: Path,
        employee_directory: EmployeeDirectory | None = None,
    ):
        self._output_dir = output_dir
        self._output_dir.mkdir(parents=True, exist_ok=True)
        self._directory = employee_directory

    def export(
        self,
        results: dict[str, list[dict]] | list[dict],
        month: str,
        plan_type: str,
        rule: dict | list | None = None,
    ) -> dict[str, Path]:
        """导出结果到 JSON 和 CSV。

        Args:
            results: skill_calculate_batch 或 skill_calculate 的输出
            month: YYYY-MM
            plan_type: 方案类型关键词
            rule: 规则 JSON（传入时额外输出薪酬提交 CSV）

        Returns:
            {"json": Path, "csv": Path} 或 {"json": Path}
        """
        ym = month.replace("-", "")
        json_path = self._export_json(results, ym, plan_type)
        paths = {"json": json_path}

        if rule is not None:
            csv_path = self._export_csv(results, rule, ym, plan_type)
            paths["csv"] = csv_path
            logger.info(f"结果导出完成: {json_path.name}, {csv_path.name}")
        else:
            logger.info(f"结果导出完成: {json_path.name}")

        return paths

    def _export_json(
        self,
        results: dict[str, list[dict]] | list[dict],
        ym: str,
        plan_type: str,
    ) -> Path:
        """写 JSON 文件。"""
        filepath = self._output_dir / f"result_{ym}_{plan_type}.json"
        filepath.write_text(
            json.dumps(results, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return filepath

    def _export_csv(
        self,
        results: dict[str, list[dict]] | list[dict],
        rule: dict | list,
        ym: str,
        plan_type: str,
    ) -> Path:
        """写薪酬提交 CSV（扁平宽表，UTF-8 with BOM）。"""
        filepath = self._output_dir / f"result_{ym}_{plan_type}.csv"

        rules = rule if isinstance(rule, list) else [rule]
        if isinstance(results, dict):
            all_rows = []
            for r in rules:
                pid = r.get("plan_id", "")
                emp_list = results.get(pid, [])
                all_rows.extend(self._build_csv_rows(r, emp_list))
        else:
            all_rows = self._build_csv_rows(rules[0], results)

        if not all_rows:
            return filepath

        headers: list[str] = []
        seen: set[str] = set()
        for row in all_rows:
            for key in row.keys():
                if key not in seen:
                    seen.add(key)
                    headers.append(key)

        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers, restval="")
            writer.writeheader()
            writer.writerows(all_rows)

        return filepath

    def _build_csv_rows(self, rule: dict, emp_list: list[dict]) -> list[dict]:
        """根据 rule + 计算结果构建 CSV 行（每员工一行）。"""
        calc_rule = rule.get("calculation_rule", {})
        method = calc_rule.get("method", "")
        max_amount = calc_rule.get("capping", {}).get("max_amount")

        rows = []
        for emp in emp_list:
            if not emp.get("eligible", False):
                continue

            display_name, employee_no = self._resolve_identity(emp.get("employee_id", ""))

            row = {
                "方案生效月份": rule.get("effective_month", ""),
                "方案名称": rule.get("plan_name", ""),
                "业务类型": rule.get("plan_type", ""),
                "方案对象": display_name,
                "员工编号": employee_no,
                "方案总金额上限": _fmt_amount(max_amount),
                "实际总金额": _fmt_amount(emp.get("amount", 0)),
            }

            detail = emp.get("detail", {})
            self._append_breakdown_columns(row, detail, method)
            rows.append(row)

        return rows

    def _resolve_identity(self, identifier) -> tuple[str, str]:
        """解析 employee_id 字段，返回 (姓名展示值, 员工编号)。

        - directory 命中：返回 (姓名, 编号)
        - directory 缺失或未命中：返回 (原值, "")
        """
        if self._directory is None:
            return ("" if identifier is None else str(identifier).strip(), "")
        return self._directory.resolve_pair(identifier)

    def _append_breakdown_columns(
        self, row: dict, detail: dict, method: str
    ) -> None:
        """按 method 类型将子项明细横向展开到 row 中。"""
        labels = "ABCDEFGHIJ"

        if method == "ratio_based":
            breakdown = detail.get("breakdown", [])
            for i, item in enumerate(breakdown):
                tag = labels[i] if i < len(labels) else str(i + 1)
                prefix = f"子项{tag}"
                row[f"{prefix}_名称"] = item.get("name", "")
                row[f"{prefix}_基数"] = _fmt_amount(item.get("base_amount", ""))
                row[f"{prefix}_目标值"] = item.get("target_value", "")
                row[f"{prefix}_实际值"] = item.get("actual", "")
                row[f"{prefix}_金额"] = _fmt_amount(item.get("raw_amount", 0))

        elif method == "rank_tiered":
            row["排名"] = detail.get("rank", "")
            row["排名指标值"] = detail.get("rank_value", "")
            row["排名奖金"] = _fmt_amount(detail.get("amount", 0))

        elif method == "tiered_commission":
            row["业绩额"] = _fmt_amount(detail.get("performance", ""))
            row["提成合计"] = _fmt_amount(detail.get("total_commission", ""))
            row["扣减合计"] = _fmt_amount(detail.get("total_deduction", ""))

        elif method == "per_unit":
            row["单价"] = _fmt_amount(detail.get("unit_price", ""))
            row["数量"] = detail.get("quantity_raw", "")
            row["有效数量"] = detail.get("quantity_capped", "")

        elif method == "fixed_bonus":
            row["条件达成"] = "是" if detail.get("condition_met") else "否"
            row["奖金"] = _fmt_amount(detail.get("amount", 0))

    def _export_excel(
        self,
        results: dict[str, list[dict]] | list[dict],
        ym: str,
        plan_type: str,
    ) -> Path:
        """写 Excel 文件，每个 plan_id 一个 sheet。"""
        filepath = self._output_dir / f"result_{ym}_{plan_type}.xlsx"
        wb = openpyxl.Workbook()

        # 删除默认 sheet（后续根据数据创建）
        default_sheet = wb.active

        # 判断是 batch 还是 single 结果
        if isinstance(results, dict):
            # batch: key = plan_id
            for plan_id, emp_list in results.items():
                ws = wb.create_sheet(title=self._safe_sheet_name(plan_id))
                self._write_result_sheet(ws, emp_list)
        else:
            # single: list of dicts
            plan_name = results[0].get("plan_name", "result") if results else "result"
            ws = wb.create_sheet(title=self._safe_sheet_name(plan_name))
            self._write_result_sheet(ws, results)

        # 删除默认空 sheet
        if default_sheet:
            wb.remove(default_sheet)

        wb.save(filepath)
        wb.close()
        return filepath

    def _write_result_sheet(self, ws, emp_list: list[dict]):
        """写一个 result sheet。"""
        headers = [
            "employee_id", "plan_id", "plan_name", "eligible",
            "amount", "raw_amount", "capped_amount", "validation_errors",
        ]

        # Header row
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Data rows
        for row_idx, emp in enumerate(emp_list, 2):
            ws.cell(row=row_idx, column=1, value=emp.get("employee_id", ""))
            ws.cell(row=row_idx, column=2, value=emp.get("plan_id", ""))
            ws.cell(row=row_idx, column=3, value=emp.get("plan_name", ""))
            ws.cell(row=row_idx, column=4, value=emp.get("eligible", ""))
            ws.cell(row=row_idx, column=5, value=emp.get("amount", 0))
            ws.cell(row=row_idx, column=6, value=emp.get("raw_amount", ""))
            ws.cell(row=row_idx, column=7, value=emp.get("capped_amount", ""))
            errors = emp.get("validation_errors", [])
            ws.cell(row=row_idx, column=8, value="; ".join(errors) if errors else "")

    def _safe_sheet_name(self, name: str) -> str:
        """Excel sheet 名最长 31 字符，且不能含特殊字符。"""
        safe = name.replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "")
        safe = safe.replace("[", "").replace("]", "").replace(":", "")
        if len(safe) > 31:
            safe = safe[:31]
        return safe
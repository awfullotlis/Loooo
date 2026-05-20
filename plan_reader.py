"""plan_reader — 方案 Excel 文件 → 结构化文本（供 rule_phrase 使用）。

方案文档是 Excel 文件，可能包含多个 sheet 代表不同激励方向。
PlanDocumentReader 将 Excel 内容转为可读文本，rule_phrase 的系统提示
已支持多方向检测，所以默认将所有 sheet 合并为一段文本。
"""

import logging
from pathlib import Path

import openpyxl

logger = logging.getLogger("incentive_bot")


class PlanDocumentReader:
    """读方案 Excel 文件并转为文本。

    默认策略：将所有有意义 sheet 合并为一段文本（带分隔标记），
    让 rule_phrase 自行判断单方向/多方向输出。
    """

    def read_plan_as_text(self, filepath: Path) -> str:
        """读方案 Excel，合并所有 sheet 为一段文本。

        Args:
            filepath: 方案 Excel 文件路径

        Returns:
            结构化文本字符串，每个 sheet 带分隔标记
        """
        wb = openpyxl.load_workbook(filepath, data_only=True)

        parts = []
        parts.append(f"=== 方案文档：{filepath.stem} ===")

        for ws in wb.worksheets:
            if not self._is_meaningful_sheet(ws):
                continue
            sheet_text = self._sheet_to_text(ws, ws.title)
            parts.append(sheet_text)

        wb.close()

        if not parts:
            logger.warning(f"方案文件 {filepath.name} 无有意义 sheet")
            return ""

        result = "\n\n".join(parts)
        logger.info(f"读取方案文件 {filepath.name}: {len(wb.worksheets)} 个 sheet")
        return result

    def read_plan_per_sheet(self, filepath: Path) -> list[str]:
        """读方案 Excel，每个 sheet 返回独立文本（高级选项）。

        适用于合并文本导致 LLM 输出质量下降的场景。
        """
        wb = openpyxl.load_workbook(filepath, data_only=True)

        texts = []
        for ws in wb.worksheets:
            if not self._is_meaningful_sheet(ws):
                continue
            texts.append(self._sheet_to_text(ws, ws.title))

        wb.close()
        return texts

    def _sheet_to_text(self, ws, sheet_name: str) -> str:
        """将单个 worksheet 转为结构化文本。

        策略：
        1. 读取所有单元格
        2. 按行分组，识别表格区域和自由文本区域
        3. 表格区域：以 tab 分隔的行
        4. 自由文本区域：原样保留
        5. Sheet 名作为标题
        """
        lines = [f"--- {sheet_name} ---"]

        rows_data = []
        for row in ws.iter_rows(values_only=True):
            cells = []
            for val in row:
                if val is not None:
                    cells.append(str(val))
                else:
                    cells.append("")
            rows_data.append(cells)

        # 找到内容区域（非全空行）
        content_rows = [r for r in rows_data if any(c.strip() for c in r)]

        if not content_rows:
            return lines[0]  # 只有标题

        # 检测是否是纯表格（首行像 header）
        # 简化策略：直接将每行转为 tab 分隔的文本
        for row_cells in content_rows:
            # 去掉尾部空单元格
            trimmed = row_cells
            while trimmed and trimmed[-1].strip() == "":
                trimmed = trimmed[:-1]
            if trimmed:
                line = "\t".join(trimmed)
                lines.append(line)

        return "\n".join(lines)

    def _is_meaningful_sheet(self, ws) -> bool:
        """判断 sheet 是否包含实际方案内容（非纯格式/空 sheet）。"""
        cell_count = 0
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if val is not None and str(val).strip():
                    cell_count += 1
        return cell_count >= 5  # 至少 5 个有内容的单元格